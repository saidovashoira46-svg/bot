import os
import asyncio
import sqlite3
from datetime import datetime, timedelta
from aiogram import Bot, Dispatcher, F
from aiogram.types import (
    Message,
    CallbackQuery,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
    FSInputFile
)
from aiogram.filters import CommandStart, Command
from openpyxl import Workbook
from openpyxl.styles import Font

# ================= CONFIG =================

TOKEN = os.getenv("BOT_TOKEN")
ADMIN_ID = int(os.getenv("ADMIN_ID", "6140962854"))

# 🔴 KANALINGNI SHU YERGA YOZ
# 🔴 2 TA KANAL
CHANNELS = [
    "@alaziz_academy",
    "@abdulaziz_avazovichY"
]


bot = Bot(token=TOKEN)
dp = Dispatcher()

# ================= DATABASE =================

conn = sqlite3.connect("votes.db")
cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS votes (
    user_id INTEGER,
    fan TEXT,
    sinf TEXT,
    student TEXT,
    UNIQUE(user_id, fan, sinf)
)
""")
conn.commit()

# ================= SUBSCRIPTION CHECK =================

async def check_subscription(user_id):
    for channel in CHANNELS:
        try:
            member = await bot.get_chat_member(channel, user_id)
            if member.status in ["left", "kicked"]:
                return False
        except:
            return False
    return True

async def force_subscribe(message_or_call):
    buttons = []

    for ch in CHANNELS:
        buttons.append([
            InlineKeyboardButton(
                text="📢 Kanalga obuna bo‘lish",
                url=f"https://t.me/{ch.replace('@','')}"
            )
        ])

    buttons.append([
        InlineKeyboardButton(text="✅ Tekshirish", callback_data="check_sub")
    ])

    await message_or_call.answer(
        "❗ Botdan foydalanish uchun kanalga obuna bo‘ling:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )

# ================= DATA =================


DATA = {
    "ENGLISH": {
        "1-6 SINF": [
            {"name": "RAZMATOV ABDULLOX", "filial": "Kids 2", "maktab": 3, "sinf": 7},
            {"name": "UMARALIYEV AXAD", "filial": "Dostobod", "maktab": 11, "sinf": 5},
            {"name": "Tolibjonov Jamshid", "filial": "Kids 1"},
            {"name": "Yo'ldosheva Hadicha", "filial": "Chinoz", "maktab": 6, "sinf": 4},
            {"name": "Muhammadiev Jamoliddin", "filial": "Xalqabod", "maktab": 56, "sinf": 4},
            {"name": "RAIMOV SHAXBOZ", "filial": "Olmazor", "maktab": 29, "sinf": 5},
            {"name": "ABDUJALLOV BUNYOD", "filial": "Gulbahor", "maktab": 1, "sinf": 5},
            {"name": "ZAKIRBAYEV SHAXBOZBEK", "filial": "Kasblar", "maktab": 6, "sinf": 4},
            {"name": "XASANOVA DIYORA", "filial": "Paxtazor", "maktab": 32, "sinf": 6},
            {"name": "EGAMBERDIYEVA GO'ZAL", "filial": "Niyazbosh", "maktab": 10, "sinf": 7},
            {"name": "TAGIROV DIYORBEK", "filial": "Mevazor", "maktab": 51, "sinf": 5},
        ],

        "7-11 SINF": [
            {"name": "ABDUQODIROVA ASAL", "filial": "Kids 2", "maktab": 3, "sinf": 8},
            {"name": "KOSIMBEKOV SULTONBEK", "filial": "Dostobod", "maktab": 11, "sinf": 10},
            {"name": "Tolibjonov Davron", "filial": "Kids 1"},
            {"name": "Murodov Azimjon", "filial": "Chinoz", "sinf": 8},
            {"name": "Davudjanova Shahzoda", "filial": "Xalqabod", "maktab": 28, "sinf": 10},
            {"name": "OCHILOVA DILNURA", "filial": "Olmazor", "maktab": 9, "sinf": 11},
            {"name": "NURIDINOV ASADBEK", "filial": "Gulbahor", "maktab": 45, "sinf": 9},
            {"name": "ABDUSATTOROV ABDUMALIK", "filial": "Kasblar", "maktab": 6, "sinf": 9},
            {"name": "HUSNIDDINOV JO'RABEK", "filial": "Paxtazor", "maktab": 5, "sinf": 11},
            {"name": "MUXIDDINOVA MUQADDAS", "filial": "Niyazbosh", "maktab": 0, "sinf": 9},
            {"name": "RASULJONOV SHOXRUX", "filial": "Mevazor", "maktab": 51, "sinf": 6},
        ]
    },
    "RUS TILI": {
        "1-6 SINF": [
            {"name": "ABDURAXMONOV UMIDJON", "filial": "Kids 2", "maktab": 5, "sinf": 3},
            {"name": "AKBARALIYEV SHOXJAHON", "filial": "Dostobod", "maktab": 24, "sinf": 6},
            {"name": "Abdurashidova Asal", "filial": "Kids 1"},
            {"name": "Qurolova Xonzoda", "filial": "Chinoz", "maktab": 18, "sinf": 6},
            {"name": "Azatov Ahmadali", "filial": "Xalqabod", "maktab": 28, "sinf": 8},
            {"name": "ARIKULOVA MUXSINA", "filial": "Olmazor", "maktab": 10, "sinf": 3},
            {"name": "MAMADIYOROV SALOXIDIN", "filial": "Gulbahor", "maktab": 40, "sinf": 6},
            {"name": "NISHANALIYEVA KUMUSHBIBI", "filial": "Kasblar", "maktab": 6, "sinf": 3},
            {"name": "XUDOYBERGANOVA ZAXRO", "filial": "Niyazbosh", "maktab": 27, "sinf": 8},
            {"name": "TURDIBOYEVA MADENA", "filial": "Mevazor", "maktab": 51, "sinf": 5},
        ],
        "7-11 SINF": [
            {"name": "RUSTAMJONOVA MUSLIMA", "filial": "Kids 2", "maktab": 42, "sinf": 8},
            {"name": "UMIRKULOVA FOTIMA", "filial": "Dostobod", "maktab": 40, "sinf": 11},
            {"name": "Asrolxonova Iroda", "filial": "Chinoz", "maktab": 26, "sinf": 10},
            {"name": "Ergashov Behrzu", "filial": "Xalqabod", "maktab": 28, "sinf": 10},
            {"name": "SOYIBJONOV JASUR", "filial": "Olmazor", "maktab": 9, "sinf": 11},
            {"name": "ERNAZOROVA MUSLIMA", "filial": "Gulbahor", "maktab": 1, "sinf": 8},
            {"name": "QODIRJONOV MUXAMMADORIF", "filial": "Kasblar", "maktab": 6, "sinf": 7},
            {"name": "SARBAYEV SUNNAT", "filial": "Paxtazor", "maktab": 50, "sinf": 7},
            {"name": "BAXTIYOROVA MUSHTARIY", "filial": "Niyazbosh", "maktab": 16, "sinf": 7},
            {"name": "TURSUNBOYEVA BAXTIGUL", "filial": "Mevazor", "maktab": 6, "sinf": 8},
        ]
    },
    "ARAB TILI": {
        "1-6 SINF": [
            {"name": "XAMIDOV MUXAMMADYUSUF", "filial": "Kids 2", "maktab": 44, "sinf": 3},
            {"name": "HURSANABODOVA MARJONA", "filial": "Dostobod", "maktab": 35, "sinf": 4},
            {"name": "Murodjonova Oisha", "filial": "Kids 1"},
            {"name": "Xadjayev Ziyovuddin", "filial": "Chinoz", "maktab": 26, "sinf": 2},
            {"name": "Azimjonov Akbar", "filial": "Xalqabod", "maktab": 12, "sinf": 4},
            {"name": "GAYRATOV SHAMSHOD", "filial": "Kasblar", "maktab": 5, "sinf": 4},
            {"name": "NABIYEV XAMIDULLOH", "filial": "Niyazbosh", "maktab": 9, "sinf": 7},
            {"name": "NURIMANOV JANDOS", "filial": "Mevazor", "maktab": 6, "sinf": 5},
        ],
        "7-11 SINF": [
            {"name": "XAMIDOV ABDURAXMON", "filial": "Kids 2", "maktab": 45, "sinf": 7},
            {"name": "USMONOVA DILORA", "filial": "Dostobod", "maktab": 10, "sinf": 7},
            {"name": "Xatamova Jasmina", "filial": "Chinoz", "maktab": 9, "sinf": 8},
            {"name": "Axmadjonova Ruxshona", "filial": "Xalqabod", "maktab": 55, "sinf": 11},
            {"name": "NISHANOVA MALIKA", "filial": "Kasblar"},
            {"name": "ERGASHBOYEVA ROBIYA", "filial": "Niyazbosh", "maktab": 27, "sinf": 9},
        ]
    },
    "HAMSHIRA": {
        "1-6 SINF": [
            {"name": "MASHOKIROVA MADINA", "filial": "Kids 2"},
            {"name": "BOXODIROVA MAXLIYO", "filial": "Olmazor", "maktab": 15, "sinf": 6},
        ],
        "7-11 SINF": [
            {"name": "TOLIPOVA NIGORA", "filial": "Kids 2"},
            {"name": "BERDENOVA TOLGONOY", "filial": "Dostobod", "maktab": 40, "sinf": 9},
            {"name": "Muhlisa Komiljonova", "filial": "Chinoz", "maktab": 36, "sinf": 11},
            {"name": "Xudoyqulova Xadicha", "filial": "Xalqabod", "maktab": 28, "sinf": 9},
            {"name": "NORMUXAMMEDOVA NILUFAR", "filial": "Olmazor", "maktab": 5, "sinf": 11},
            {"name": "TURSUNBOYEVA RAYXONA", "filial": "Kasblar", "maktab": 1, "sinf": 11},
            {"name": "ASQARXO'JAYEVA SEVINCH", "filial": "Paxtazor", "maktab": 11, "sinf": 10},
        ]
    },
    "IT": {

    "1-6 SINF": [
        {"name": "Ahmadjonov Mahmudjon", "filial": "Chinoz", "maktab": 26, "sinf": 7},
        {"name": "Xolqo'ziyev Asadbek", "filial": "Xalqabod", "maktab": 28, "sinf": 6},
        {"name": "DONYOROV MUXAMMADALI", "filial": "Olmazor", "sinf": 7},
        {"name": "XAYTBOYEV OTABEK", "filial": "Gulbahor", "maktab": 1, "sinf": 9},
        {"name": "RASULOV DAVRON", "filial": "Kasblar", "maktab": 10, "sinf": 6},
    ],

    "7-11 SINF": [
        {"name": "Mengbayev Ma'mur", "filial": "Chinoz", "maktab": 8, "sinf": 11},
        {"name": "Jalolov Mirasad", "filial": "Xalqabod", "maktab": 56, "sinf": 7},
        {"name": "SULTONMURODOV BEXRUZ", "filial": "Olmazor", "maktab": 15, "sinf": 10},
        {"name": "TO'LQINOVA MARJONA", "filial": "Gulbahor", "maktab": 48, "sinf": 8},
        {"name": "JABBOROV JAVOXIR", "filial": "Kasblar", "maktab": 6, "sinf": 6},
    ]
    },
    "KOMPYUTER": {
        "1-6 SINF": [
            {"name": "Izatullayev Sarvar", "filial": "Chinoz", "maktab": 40, "sinf": 5},
            {"name": "Zulpiyev Otabek", "filial": "Olmazor", "maktab": 27, "sinf": 6},
            {"name": "QURBONALIYEV MUSLIMBEK", "filial": "Kasblar", "maktab": 27, "sinf": 3},
        ],

        "7-11 SINF": [
            {"name": "G'aniyev Kamron", "filial": "Chinoz", "maktab": 40, "sinf": 11},
            {"name": "Abduraxmonova Sevinch", "filial": "Olmazor", "maktab": 9, "sinf": 10},
            {"name": "NAVJABAYEVA NILUFAR", "filial": "Kasblar"},
        ]
    },
    "MATEMATIKA": {
        "1-6 SINF": [
            {"name": "FAXRIDDINOV FARXOD", "filial": "Dostobod", "maktab": 11, "sinf": 5},
            {"name": "Isroildjonova Zohida", "filial": "Chinoz", "maktab": 21, "sinf": 5},
            {"name": "QURBONOVA DIANA", "filial": "Olmazor", "maktab": 5, "sinf": 5},
            {"name": "ROZIMURODOVA SHAXNOZA", "filial": "Kasblar", "maktab": 9, "sinf": 3},
            {"name": "SHAVKATOV ABRORJON", "filial": "Niyazbosh", "maktab": 16, "sinf": 4},
        ],
        "7-11 SINF": [
            {"name": "RAXIMBERDIYEV FAZLIDDIN", "filial": "Dostobod", "maktab": 29, "sinf": 9},
            {"name": "Eshmatov Muhammadrizo", "filial": "Chinoz", "maktab": 4, "sinf": 11},
            {"name": "ABDUJALOLOV JALOLIDDIN", "filial": "Olmazor", "maktab": 17, "sinf": 8},
            {"name": "MAXMUDJONOVA MUSHTARIY", "filial": "Kasblar", "maktab": 6, "sinf": 8},
            {"name": "NIGMATULLAYEVA MUSLIMA", "filial": "Niyazbosh", "maktab": 27, "sinf": 7},
            {"name": "ABDIXAFIZOV NURBOL", "filial": "Mevazor", "maktab": 36, "sinf": 7},
        ]
    },
    "KOREYS TILI": {
        "1-6 SINF": [
            {"name": "DJORAYEVA FARIZA", "filial": "Olmazor", "maktab": 15, "sinf": 5},
            {"name": "MAXKAMBOYEVA MADINA", "filial": "Niyazbosh", "maktab": 13, "sinf": 11},
        ],
        "7-11 SINF": [
            {"name": "TOQIBOYEVA ZIYODA", "filial": "Dostobod", "maktab": 19, "sinf": 9},
            {"name": "Bolyeva Charosxon", "filial": "Chinoz", "maktab": 38, "sinf": 8},
            {"name": "TUREJANOVA AIDA", "filial": "Olmazor", "maktab": 42, "sinf": 11},
            {"name": "NISHANBAYEVA LAYLO", "filial": "Kasblar"},
            {"name": "ABDULLAYEVA ZARINA", "filial": "Paxtazor", "maktab": 33, "sinf": 10},
            {"name": "ABDUSHAKIROVA YASMINA", "filial": "Niyazbosh", "maktab": 13, "sinf": 11},
        ]
    },
    "MENTAL": {
        "1-6 SINF": [
            {"name": "QODIROV ULUG'BEK", "filial": "Kids 2", "maktab": 12, "sinf": 2},
            {"name": "ISAYEV XUSAN", "filial": "Olmazor", "maktab": 5, "sinf": 2},
        ],
        "7-11 SINF": [
            {"name": "XAMIDJONOV ABDURAXMON", "filial": "Kids 2", "maktab": 3, "sinf": 6},
            {"name": "SIDDIQOVA INTIZOR", "filial": "Olmazor", "maktab": 10, "sinf": 1},
        ]
    },
    "LOYIHA": {
        "1-6 SINF": [
            {"name": "Sheraliyev Ismoil", "filial": "Kids 1", "sinf": 1},
            {"name": "Botirjonova Malika", "filial": "Kids 1", "sinf": 2},
            {"name": "Artikaliyev Boburjon", "filial": "Kids 1", "sinf": 2},
            {"name": "Nazarmatova Iymona", "filial": "Kids 1", "sinf": 3},
            {"name": "Shahramboyeva Oisha", "filial": "Kids 1", "sinf": 3},
            {"name": "Shuhratova Muslima", "filial": "Kids 1", "sinf": 4},
            {"name": "Tursunaliyev Nodirbek", "filial": "Kids 1", "sinf": 4},
        ],
        "7-11 SINF": []
    },
}


# ================= VOTING =================

VOTING_DURATION_MINUTES = 3
START_TIME = datetime.now()


def voting_active():
    return datetime.now() < START_TIME + timedelta(minutes=VOTING_DURATION_MINUTES)


def get_remaining_time():
    end_time = START_TIME + timedelta(minutes=VOTING_DURATION_MINUTES)
    remaining = end_time - datetime.now()

    if remaining.total_seconds() <= 0:
        return "0:00"

    minutes = int(remaining.total_seconds() // 60)
    seconds = int(remaining.total_seconds() % 60)

    return f"{minutes}:{seconds:02d}"


# ================= DATABASE FUNCS =================

def add_vote(user_id, fan, sinf, student):
    try:
        cursor.execute(
            "INSERT INTO votes VALUES (?, ?, ?, ?)",
            (user_id, fan, sinf, student)
        )
        conn.commit()
        return True
    except:
        return False

def get_results(fan, sinf):
    cursor.execute("""
        SELECT student, COUNT(*)
        FROM votes
        WHERE fan=? AND sinf=?
        GROUP BY student
    """, (fan, sinf))
    return cursor.fetchall()

# ================= MENU =================
def get_results_text(fan, sinf):
    results = get_results(fan, sinf)
    result_dict = {student: count for student, count in results}
    total = sum(result_dict.values())

    text = f"\n📊 {fan} — {sinf}\n\n"

    for student in DATA[fan][sinf]:
        name = student["name"]
        filial = student.get("filial", "")
        count = result_dict.get(name, 0)
        percent = (count / total * 100) if total else 0

        text += f"{name} ({filial}) — {count} ta ({percent:.1f}%)\n"

    text += f"\n🗳 Jami ovoz: {total}\n"
    return text

def generate_excel():
    wb = Workbook()
    wb.remove(wb.active)

    for fan in DATA:
        ws = wb.create_sheet(title=fan[:31])  # Sheet nomi 31 belgidan oshmasin

        ws.append(["Sinf", "O‘quvchi", "Filial", "Ovoz", "Foiz"])

        # Header bold
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for sinf in DATA[fan]:
            results = get_results(fan, sinf)
            result_dict = {student: count for student, count in results}
            total = sum(result_dict.values())

            for student in DATA[fan][sinf]:
                name = student["name"]
                filial = student.get("filial", "")
                count = result_dict.get(name, 0)
                percent = (count / total * 100) if total else 0

                ws.append([
                    sinf,
                    name,
                    filial,
                    count,
                    round(percent, 2)
                ])

            # Sinflar orasiga bo‘sh qatordan ajratish
            ws.append([])

    filename = "natijalar.xlsx"
    wb.save(filename)
    return filename
@dp.message(Command("excel"))
async def send_excel(message: Message):
    if message.from_user.id != ADMIN_ID:
        await message.answer("❌ Siz admin emassiz.")
        return

    filename = generate_excel()

    await message.answer_document(
        FSInputFile(filename),
        caption="📊 Barcha natijalar (Excel)"
    )

@dp.callback_query(F.data == "results")
async def results_handler(call: CallbackQuery):
    await call.answer()

    buttons = [
        [InlineKeyboardButton(text=fan, callback_data=f"resfan|{fan}")]
        for fan in DATA
    ]

    await call.message.edit_text(
        "📊 Qaysi fan natijasini ko‘rmoqchisiz?",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )
@dp.callback_query(F.data.startswith("resfan|"))
async def results_fan(call: CallbackQuery):
    await call.answer()
    fan = call.data.split("|")[1]

    buttons = [
        [InlineKeyboardButton(text=sinf, callback_data=f"ressinf|{fan}|{sinf}")]
        for sinf in DATA[fan]
    ]

    await call.message.edit_text(
        f"📊 {fan}\nSinf tanlang:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )

@dp.callback_query(F.data.startswith("ressinf|"))
async def results_sinf(call: CallbackQuery):
    await call.answer()

    _, fan, sinf = call.data.split("|")

    await call.message.edit_text(get_results_text(fan, sinf))

async def show_menu(message):
    buttons = []

    if voting_active():
        for fan in DATA:
            buttons.append(
                [InlineKeyboardButton(text=fan, callback_data=f"fan|{fan}")]
            )

    buttons.append(
        [InlineKeyboardButton(text="📊 Natijalar", callback_data="results")]
    )

    await message.answer(
        "📚 Fan tanlang:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )

# ================= START =================

@dp.message(CommandStart())
async def start_handler(message: Message):

    subscribed = await check_subscription(message.from_user.id)

    if not subscribed:
        await force_subscribe(message)
        return

    if voting_active():
        await message.answer(f"⏳ Qolgan vaqt: {get_remaining_time()}")
    else:
        await message.answer("❌ Ovoz berish tugagan.")

    await show_menu(message)

# ================= CHECK BUTTON =================

@dp.callback_query(F.data == "check_sub")
async def check_sub_handler(call: CallbackQuery):
    await call.answer()

    subscribed = await check_subscription(call.from_user.id)

    if not subscribed:
        await call.message.answer("❌ Hali ham obuna bo‘lmagansiz.")
        return

    await call.message.answer("✅ Obuna tasdiqlandi!")
    await show_menu(call.message)

# ================= FAN =================

@dp.callback_query(F.data.startswith("fan|"))
async def fan_handler(call: CallbackQuery):
    await call.answer()

    if not await check_subscription(call.from_user.id):
        await force_subscribe(call.message)
        return

    fan = call.data.split("|")[1]

    buttons = [
        [InlineKeyboardButton(text=sinf, callback_data=f"sinf|{fan}|{sinf}")]
        for sinf in DATA[fan]
    ]

    await call.message.edit_text(
        "🏫 Sinf tanlang:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )

# ================= SINF =================

@dp.callback_query(F.data.startswith("sinf|"))
async def sinf_handler(call: CallbackQuery):
    await call.answer()

    if not await check_subscription(call.from_user.id):
        await force_subscribe(call.message)
        return

    _, fan, sinf = call.data.split("|")

    buttons = []

    for student in DATA[fan][sinf]:
        name = student["name"]
        filial = student.get("filial", "")
        maktab = student.get("maktab", None)

        text = f"{name}"

        if filial:
            text += f" ({filial})"

        if maktab:
            text += f" {maktab}-m"

        buttons.append(
            [InlineKeyboardButton(
                text=text,
                callback_data=f"vote|{fan}|{sinf}|{name}"
            )]
        )


    await call.message.edit_text(
        "👨‍🎓 O‘quvchini tanlang:",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=buttons)
    )

# ================= VOTE =================

@dp.callback_query(F.data.startswith("vote|"))
async def vote_handler(call: CallbackQuery):
    await call.answer()

    if not await check_subscription(call.from_user.id):
        await force_subscribe(call.message)
        return

    _, fan, sinf, student = call.data.split("|")

    if not voting_active():
        await call.message.answer("❌ Ovoz berish tugagan.")
        return

    success = add_vote(call.from_user.id, fan, sinf, student)

    if not success:
        await call.message.answer("❌ Siz allaqachon ovoz bergansiz!")
        return

    await call.message.answer("✅ Ovozingiz qabul qilindi!")

# ================= RUN =================

async def main():
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
